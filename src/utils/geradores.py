import tempfile
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.units import inch
from weasyprint import HTML, CSS
from io import StringIO

def gerar_excel_avancado(relatorio):
    """Gera um arquivo Excel avançado com gráficos e formatação"""
    
    # Criar workbook
    wb = Workbook()
    
    # Remover planilha padrão
    wb.remove(wb.active)
    
    # === PLANILHA 1: RESUMO EXECUTIVO ===
    ws_resumo = wb.create_sheet("Resumo Executivo")
    
    # Cabeçalho
    ws_resumo['A1'] = "RELATÓRIO DE SATISFAÇÃO - TRANSPORTE MUNICIPAL"
    ws_resumo['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_resumo['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_resumo['A1'].alignment = Alignment(horizontal="center")
    ws_resumo.merge_cells('A1:F1')
    
    # Informações básicas
    row = 3
    info_data = [
        ("Linha:", relatorio.linha_numero),
        ("Período:", f"{relatorio.periodo_inicio.strftime('%d/%m/%Y')} a {relatorio.periodo_fim.strftime('%d/%m/%Y')}"),
        ("Total de Pesquisas:", relatorio.total_pesquisas),
        ("Data do Relatório:", relatorio.data_criacao.strftime('%d/%m/%Y às %H:%M')),
        ("Média Geral:", f"{relatorio.media_geral:.1f}/10 ({relatorio.get_classificacao_geral()})")
    ]
    
    for label, value in info_data:
        ws_resumo[f'A{row}'] = label
        ws_resumo[f'A{row}'].font = Font(bold=True)
        ws_resumo[f'B{row}'] = value
        row += 1
    
    # Médias por categoria
    row += 2
    ws_resumo[f'A{row}'] = "MÉDIAS POR CATEGORIA"
    ws_resumo[f'A{row}'].font = Font(size=14, bold=True)
    ws_resumo[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws_resumo.merge_cells(f'A{row}:D{row}')
    
    row += 1
    headers = ["Categoria", "Média", "Classificação", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws_resumo.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    categorias_data = [
        ("Pontualidade", relatorio.media_pontualidade),
        ("Frequência", relatorio.media_frequencia),
        ("Conforto", relatorio.media_conforto),
        ("Atendimento", relatorio.media_atendimento),
        ("Infraestrutura", relatorio.media_infraestrutura)
    ]
    
    for categoria, media in categorias_data:
        row += 1
        classificacao = relatorio.classificar_nota(media)
        status = "✅" if media >= 7 else "⚠️" if media >= 4 else "❌"
        
        ws_resumo[f'A{row}'] = categoria
        ws_resumo[f'B{row}'] = f"{media:.1f}"
        ws_resumo[f'C{row}'] = classificacao
        ws_resumo[f'D{row}'] = status
        
        # Colorir baseado na classificação
        if media >= 7:
            fill_color = "C6EFCE"  # Verde claro
        elif media >= 4:
            fill_color = "FFEB9C"  # Amarelo claro
        else:
            fill_color = "FFC7CE"  # Vermelho claro
        
        for col in range(1, 5):
            ws_resumo.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Ajustar largura das colunas
    ws_resumo.column_dimensions['A'].width = 20
    ws_resumo.column_dimensions['B'].width = 10
    ws_resumo.column_dimensions['C'].width = 15
    ws_resumo.column_dimensions['D'].width = 10
    
    # === PLANILHA 2: DADOS DETALHADOS ===
    ws_dados = wb.create_sheet("Dados Detalhados")
    
    # Cabeçalho
    ws_dados['A1'] = "DADOS DETALHADOS DAS PESQUISAS"
    ws_dados['A1'].font = Font(size=14, bold=True)
    ws_dados['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_dados['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_dados.merge_cells('A1:I1')
    
    # Headers da tabela
    headers_dados = ["ID", "Data", "Itinerário", "Pontualidade", "Frequência", "Conforto", "Atendimento", "Infraestrutura", "Observações"]
    for col, header in enumerate(headers_dados, 1):
        cell = ws_dados.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Dados das pesquisas
    pesquisas = relatorio.get_dados_pesquisas()
    for row_idx, pesquisa in enumerate(pesquisas, 3):
        ws_dados[f'A{row_idx}'] = pesquisa['id']
        ws_dados[f'B{row_idx}'] = datetime.fromisoformat(pesquisa['data_criacao']).strftime('%d/%m/%Y %H:%M')
        ws_dados[f'C{row_idx}'] = pesquisa['linha_itinerario'] or ''
        ws_dados[f'D{row_idx}'] = pesquisa['pontualidade']
        ws_dados[f'E{row_idx}'] = pesquisa['frequencia']
        ws_dados[f'F{row_idx}'] = pesquisa['conforto']
        ws_dados[f'G{row_idx}'] = pesquisa['atendimento']
        ws_dados[f'H{row_idx}'] = pesquisa['infraestrutura']
        ws_dados[f'I{row_idx}'] = pesquisa['observacoes'] or ''
    
    # Ajustar largura das colunas
    column_widths = [8, 18, 25, 12, 12, 12, 12, 15, 40]
    for col, width in enumerate(column_widths, 1):
        ws_dados.column_dimensions[chr(64 + col)].width = width
    
    # === PLANILHA 3: GRÁFICOS ===
    ws_graficos = wb.create_sheet("Gráficos")
    
    # Dados para o gráfico
    ws_graficos['A1'] = "Categoria"
    ws_graficos['B1'] = "Média"
    
    for row_idx, (categoria, media) in enumerate(categorias_data, 2):
        ws_graficos[f'A{row_idx}'] = categoria
        ws_graficos[f'B{row_idx}'] = media
    
    # Criar gráfico de barras
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"Médias por Categoria - Linha {relatorio.linha_numero}"
    chart.y_axis.title = 'Média (0-10)'
    chart.x_axis.title = 'Categorias'
    
    data = Reference(ws_graficos, min_col=2, min_row=1, max_row=6, max_col=2)
    cats = Reference(ws_graficos, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws_graficos.add_chart(chart, "D2")
    
    # === PLANILHA 4: OBSERVAÇÕES ===
    ws_obs = wb.create_sheet("Observações")
    
    ws_obs['A1'] = "OBSERVAÇÕES DOS USUÁRIOS"
    ws_obs['A1'].font = Font(size=14, bold=True)
    ws_obs['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_obs['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_obs.merge_cells('A1:B1')
    
    observacoes = relatorio.get_observacoes_lista()
    if observacoes:
        ws_obs['A2'] = "Nº"
        ws_obs['B2'] = "Observação"
        for col in ['A', 'B']:
            ws_obs[f'{col}2'].font = Font(bold=True)
            ws_obs[f'{col}2'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for idx, obs in enumerate(observacoes, 1):
            row = idx + 2
            ws_obs[f'A{row}'] = idx
            ws_obs[f'B{row}'] = obs
            ws_obs[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
    else:
        ws_obs['A3'] = "Nenhuma observação foi fornecida neste período."
        ws_obs['A3'].font = Font(italic=True)
    
    ws_obs.column_dimensions['A'].width = 5
    ws_obs.column_dimensions['B'].width = 80
    
    # === PLANILHA 5: RECOMENDAÇÕES ===
    ws_rec = wb.create_sheet("Recomendações")
    
    ws_rec['A1'] = "RECOMENDAÇÕES BASEADAS NAS AVALIAÇÕES"
    ws_rec['A1'].font = Font(size=14, bold=True)
    ws_rec['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_rec['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_rec.merge_cells('A1:B1')
    
    recomendacoes = relatorio.get_recomendacoes()
    for idx, rec in enumerate(recomendacoes, 1):
        row = idx + 2
        ws_rec[f'A{row}'] = f"{idx}."
        ws_rec[f'B{row}'] = rec
        ws_rec[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
        
        # Colorir recomendações críticas
        if "🔴" in rec:
            fill_color = "FFC7CE"  # Vermelho claro
        elif "✅" in rec:
            fill_color = "C6EFCE"  # Verde claro
        else:
            fill_color = "FFFFFF"  # Branco
        
        for col in ['A', 'B']:
            ws_rec[f'{col}{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    ws_rec.column_dimensions['A'].width = 5
    ws_rec.column_dimensions['B'].width = 80
    
    # Salvar arquivo temporário
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        return tmp.name

def gerar_pdf_avancado(relatorio):
    """Gera um PDF avançado usando ReportLab"""
    
    # Criar arquivo temporário
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
        doc = SimpleDocTemplate(tmp.name, pagesize=A4)
        story = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # Centralizado
            textColor=colors.darkblue
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        
        # Título
        story.append(Paragraph("🚌 RELATÓRIO DE SATISFAÇÃO", title_style))
        story.append(Paragraph("Sistema Municipal de Transporte Coletivo", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Informações básicas
        info_data = [
            ['Linha:', relatorio.linha_numero],
            ['Período:', f"{relatorio.periodo_inicio.strftime('%d/%m/%Y')} a {relatorio.periodo_fim.strftime('%d/%m/%Y')}"],
            ['Total de Pesquisas:', str(relatorio.total_pesquisas)],
            ['Média Geral:', f"{relatorio.media_geral:.1f}/10 ({relatorio.get_classificacao_geral()})"],
            ['Data do Relatório:', relatorio.data_criacao.strftime('%d/%m/%Y às %H:%M')]
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Médias por categoria
        story.append(Paragraph("📈 Análise Detalhada por Categoria", heading_style))
        
        categorias_data = [
            ['Categoria', 'Média', 'Classificação'],
            ['🕐 Pontualidade', f"{relatorio.media_pontualidade:.1f}", relatorio.classificar_nota(relatorio.media_pontualidade)],
            ['⏱️ Frequência', f"{relatorio.media_frequencia:.1f}", relatorio.classificar_nota(relatorio.media_frequencia)],
            ['🚌 Conforto', f"{relatorio.media_conforto:.1f}", relatorio.classificar_nota(relatorio.media_conforto)],
            ['👥 Atendimento', f"{relatorio.media_atendimento:.1f}", relatorio.classificar_nota(relatorio.media_atendimento)],
            ['🏢 Infraestrutura', f"{relatorio.media_infraestrutura:.1f}", relatorio.classificar_nota(relatorio.media_infraestrutura)]
        ]
        
        categorias_table = Table(categorias_data, colWidths=[2.5*inch, 1*inch, 1.5*inch])
        categorias_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        story.append(categorias_table)
        story.append(Spacer(1, 20))
        
        # Observações
        observacoes = relatorio.get_observacoes_lista()
        if observacoes:
            story.append(Paragraph("💬 Observações dos Usuários", heading_style))
            for idx, obs in enumerate(observacoes, 1):
                story.append(Paragraph(f"<b>{idx}.</b> {obs}", styles['Normal']))
                story.append(Spacer(1, 6))
        
        story.append(Spacer(1, 20))
        
        # Recomendações
        story.append(Paragraph("📋 Recomendações", heading_style))
        recomendacoes = relatorio.get_recomendacoes()
        for rec in recomendacoes:
            story.append(Paragraph(f"• {rec}", styles['Normal']))
            story.append(Spacer(1, 6))
        
        # Rodapé
        story.append(Spacer(1, 30))
        story.append(Paragraph("___", styles['Normal']))
        story.append(Paragraph("📧 Relatório gerado automaticamente pelo Sistema de Pesquisa de Satisfação", styles['Normal']))
        story.append(Paragraph("🌐 Sistema Municipal de Transporte Coletivo", styles['Normal']))
        
        # Construir PDF
        doc.build(story)
        return tmp.name

def gerar_word_avancado(relatorio):
    """Gera um documento Word usando HTML e WeasyPrint"""
    
    html_content = f"""
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4;
                margin: 2cm;
            }}
            body {{
                font-family: 'Times New Roman', serif;
                line-height: 1.6;
                color: #333;
            }}
            .header {{
                text-align: center;
                border-bottom: 3px solid #4472C4;
                padding-bottom: 20px;
                margin-bottom: 30px;
            }}
            .title {{
                font-size: 24px;
                font-weight: bold;
                color: #4472C4;
                margin-bottom: 10px;
            }}
            .subtitle {{
                font-size: 16px;
                color: #666;
            }}
            .info-table {{
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
            }}
            .info-table th, .info-table td {{
                border: 1px solid #ddd;
                padding: 12px;
                text-align: left;
            }}
            .info-table th {{
                background-color: #f8f9fa;
                font-weight: bold;
            }}
            .section-title {{
                font-size: 18px;
                font-weight: bold;
                color: #4472C4;
                margin: 30px 0 15px 0;
                border-bottom: 2px solid #4472C4;
                padding-bottom: 5px;
            }}
            .metric {{
                margin: 15px 0;
                padding: 15px;
                border-left: 5px solid #4472C4;
                background-color: #f8f9fa;
            }}
            .metric.excellent {{ border-left-color: #28a745; background-color: #d4edda; }}
            .metric.good {{ border-left-color: #17a2b8; background-color: #d1ecf1; }}
            .metric.regular {{ border-left-color: #ffc107; background-color: #fff3cd; }}
            .metric.poor {{ border-left-color: #dc3545; background-color: #f8d7da; }}
            .observation {{
                margin: 10px 0;
                padding: 10px;
                background-color: #f8f9fa;
                border-left: 3px solid #4472C4;
            }}
            .recommendation {{
                margin: 10px 0;
                padding: 10px;
                background-color: #fff3cd;
                border-left: 3px solid #ffc107;
            }}
            .footer {{
                margin-top: 50px;
                padding-top: 20px;
                border-top: 1px solid #ddd;
                text-align: center;
                font-size: 12px;
                color: #666;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">🚌 RELATÓRIO DE SATISFAÇÃO</div>
            <div class="subtitle">Sistema Municipal de Transporte Coletivo</div>
            <p><strong>Linha:</strong> {relatorio.linha_numero}</p>
            <p>Relatório baseado em {relatorio.total_pesquisas} pesquisas</p>
        </div>
        
        <table class="info-table">
            <tr><th>Período</th><td>{relatorio.periodo_inicio.strftime('%d/%m/%Y')} a {relatorio.periodo_fim.strftime('%d/%m/%Y')}</td></tr>
            <tr><th>Total de Pesquisas</th><td>{relatorio.total_pesquisas}</td></tr>
            <tr><th>Média Geral</th><td>{relatorio.media_geral:.1f}/10 ({relatorio.get_classificacao_geral()})</td></tr>
            <tr><th>Data do Relatório</th><td>{relatorio.data_criacao.strftime('%d/%m/%Y às %H:%M')}</td></tr>
        </table>
        
        <div class="section-title">📈 Análise Detalhada por Categoria</div>
        
        <div class="metric {relatorio.classificar_nota(relatorio.media_pontualidade).lower()}">
            <h3>🕐 1. Pontualidade e Cumprimento de Horários</h3>
            <p><strong>Nota:</strong> {relatorio.media_pontualidade:.1f}/10 - {relatorio.classificar_nota(relatorio.media_pontualidade)}</p>
            <p>Avalia a confiabilidade do sistema em relação aos horários divulgados.</p>
        </div>
        
        <div class="metric {relatorio.classificar_nota(relatorio.media_frequencia).lower()}">
            <h3>⏱️ 2. Frequência e Intervalo entre Ônibus</h3>
            <p><strong>Nota:</strong> {relatorio.media_frequencia:.1f}/10 - {relatorio.classificar_nota(relatorio.media_frequencia)}</p>
            <p>Mede a adequação da oferta de serviço e tempo de espera.</p>
        </div>
        
        <div class="metric {relatorio.classificar_nota(relatorio.media_conforto).lower()}">
            <h3>🚌 3. Conforto e Condições dos Veículos</h3>
            <p><strong>Nota:</strong> {relatorio.media_conforto:.1f}/10 - {relatorio.classificar_nota(relatorio.media_conforto)}</p>
            <p>Analisa a qualidade da frota, limpeza e condições gerais.</p>
        </div>
        
        <div class="metric {relatorio.classificar_nota(relatorio.media_atendimento).lower()}">
            <h3>👥 4. Qualidade do Atendimento</h3>
            <p><strong>Nota:</strong> {relatorio.media_atendimento:.1f}/10 - {relatorio.classificar_nota(relatorio.media_atendimento)}</p>
            <p>Avalia o fator humano: motoristas e cobradores.</p>
        </div>
        
        <div class="metric {relatorio.classificar_nota(relatorio.media_infraestrutura).lower()}">
            <h3>🏢 5. Infraestrutura dos Pontos e Terminais</h3>
            <p><strong>Nota:</strong> {relatorio.media_infraestrutura:.1f}/10 - {relatorio.classificar_nota(relatorio.media_infraestrutura)}</p>
            <p>Examina as condições de espera, cobertura, bancos e segurança.</p>
        </div>
    """
    
    # Adicionar observações
    observacoes = relatorio.get_observacoes_lista()
    if observacoes:
        html_content += f"""
        <div class="section-title">💬 Observações dos Usuários ({len(observacoes)} comentários)</div>
        """
        for idx, obs in enumerate(observacoes, 1):
            html_content += f'<div class="observation"><strong>#{idx}:</strong> {obs}</div>'
    else:
        html_content += """
        <div class="section-title">💬 Observações dos Usuários</div>
        <p style="font-style: italic; color: #666;">Nenhuma observação adicional foi fornecida neste período.</p>
        """
    
    # Adicionar recomendações
    recomendacoes = relatorio.get_recomendacoes()
    html_content += """
    <div class="section-title">📋 Recomendações</div>
    """
    for rec in recomendacoes:
        html_content += f'<div class="recommendation">{rec}</div>'
    
    html_content += f"""
        <div class="footer">
            <p>📧 Relatório gerado automaticamente pelo Sistema de Pesquisa de Satisfação</p>
            <p>🌐 Sistema Municipal de Transporte Coletivo | Gerado em {relatorio.data_criacao.strftime('%d/%m/%Y às %H:%M')}</p>
        </div>
    </body>
    </html>
    """
    
    # Gerar PDF usando WeasyPrint (que pode ser convertido para Word)
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
        HTML(string=html_content).write_pdf(tmp.name)
        return tmp.name

